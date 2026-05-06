import pandas as pd
import requests
from io import StringIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


TWSE_URL = "https://mopsfin.twse.com.tw/opendata/t187ap03_L.csv"
TPEX_URL = "https://mopsfin.twse.com.tw/opendata/t187ap03_O.csv"
TWSE_ETF_URL = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
TPEX_ETF_URL = "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes"

def read_mops_csv(url: str) -> pd.DataFrame:
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()

    for enc in ["utf-8-sig", "big5-hkscs", "big5"]:
        try:
            text = resp.content.decode(enc)
            return pd.read_csv(StringIO(text))
        except Exception:
            continue

    raise RuntimeError(f"無法解碼 CSV: {url}")


def normalize(df: pd.DataFrame, market: str, suffix: str) -> pd.DataFrame:
    required_cols = ["公司代號", "公司簡稱", "產業別"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"{market} 資料缺少欄位: {missing}")

    out = pd.DataFrame()
    out["code"] = df["公司代號"].astype(str).str.strip()
    out["name"] = df["公司簡稱"].astype(str).str.strip()
    out["market"] = market
    out["industry"] = df["產業別"].astype(str).str.strip()
    out["symbol"] = out["code"] + suffix

    out["llm_input"] = (
        out["symbol"]
        + " "
        + out["name"]
        + "；市場："
        + out["market"]
        + "；官方產業別："
        + out["industry"]
    )

    return out[["symbol", "name", "market", "industry", "llm_input"]]

def is_etf_code(code: str) -> bool:
    code = str(code).strip().upper()
    return code.startswith("00") and not code.startswith("020")


def normalize_etf(rows, market: str, suffix: str) -> pd.DataFrame:
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["symbol", "name", "market", "industry", "llm_input"])

    df["code"] = df["code"].astype(str).str.strip().str.upper()
    df["name"] = df["name"].astype(str).str.strip()
    df = df[df["code"].apply(is_etf_code)].copy()

    df["symbol"] = df["code"] + suffix
    df["market"] = market
    df["industry"] = "ETF"
    df["llm_input"] = (
        df["symbol"]
        + " "
        + df["name"]
        + "；市場："
        + df["market"]
        + "；官方產業別：ETF"
    )

    return df[["symbol", "name", "market", "industry", "llm_input"]]


def get_twse_etf() -> pd.DataFrame:
    data = requests.get(TWSE_ETF_URL, timeout=30).json()

    rows = []
    for r in data:
        code = r.get("Code") or r.get("證券代號")
        name = r.get("Name") or r.get("證券名稱")
        if code and name:
            rows.append({"code": code, "name": name})

    return normalize_etf(rows, market="上市", suffix=".TW")


def get_tpex_etf() -> pd.DataFrame:
    data = requests.get(TPEX_ETF_URL, timeout=30).json()

    rows = []
    for r in data:
        code = (
            r.get("SecuritiesCompanyCode")
            or r.get("Code")
            or r.get("證券代號")
        )
        name = (
            r.get("CompanyName")
            or r.get("Name")
            or r.get("證券名稱")
        )
        if code and name:
            rows.append({"code": code, "name": name})

    return normalize_etf(rows, market="上櫃", suffix=".TWO")


def format_excel(path: str):
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "tw_stock_llm_source"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 14,  # symbol
        "B": 18,  # name
        "C": 10,  # market
        "D": 18,  # industry
        "E": 70,  # llm_input
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    for cell in ws["E"]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def main():
    listed = normalize(read_mops_csv(TWSE_URL), market="上市", suffix=".TW")
    otc = normalize(read_mops_csv(TPEX_URL), market="上櫃", suffix=".TWO")

    twse_etf = get_twse_etf()
    tpex_etf = get_tpex_etf()
    
    df = pd.concat([listed, otc, twse_etf, tpex_etf], ignore_index=True)
    df = df.drop_duplicates(subset=["symbol"])
    df = df.sort_values("symbol")

    xlsx_path = "tw_stock_llm_source.xlsx"
    csv_path = "tw_stock_llm_source.csv"

    df.to_excel(xlsx_path, index=False, engine="openpyxl")
    format_excel(xlsx_path)

    # 備用 CSV，若你的批次框架比較適合吃 CSV，也可以直接用
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    print(f"完成：{xlsx_path}，共 {len(df)} 筆")
    print(f"備用輸出：{csv_path}")


if __name__ == "__main__":
    main()
